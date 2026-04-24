from pathlib import Path
from openpyxl import load_workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
import pandas as pd  # type: ignore


def set_column_widths_mm(
    xlsx: Path,
    default_mm: float,
    special: dict[int, float],
) -> None:
    mm_to_excel = 1.92

    wb = load_workbook(xlsx)
    ws = wb.active

    for i in range(1, ws.max_column + 1):
        width_mm = special.get(i, default_mm)
        ws.column_dimensions[get_column_letter(i)].width = width_mm / mm_to_excel

    wb.save(xlsx)


def freeze_header_and_first_col(xlsx: Path) -> None:
    wb = load_workbook(xlsx)
    ws = wb.active

    # Freeze everything above row 2 and left of column B
    ws.freeze_panes = "B2"

    wb.save(xlsx)


def write_xlsx_formatted(
    df: pd.DataFrame,
    path: Path,
    index: bool = True,
    col_widths_mm: float = 10.0,
    special_widths_mm: dict[int, float] = {},
    freeze: tuple[int, int] = (1, 1),
) -> None:
    mm_to_excel = 1.92

    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=index, sheet_name="Sheet1")

        wb = writer.book
        ws = writer.sheets["Sheet1"]

        # Column widths
        for i, _ in enumerate(df.columns):
            ws.set_column(i, i, col_widths_mm / mm_to_excel)
        for col_idx, width_mm in special_widths_mm.items():
            ws.set_column(col_idx, col_idx, width_mm / mm_to_excel)

        # Freeze header + first column
        ws.freeze_panes(freeze[0], freeze[1])
        del wb
